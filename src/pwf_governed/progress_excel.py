"""Human-readable Project Progress Excel generator and preservation engine.

Provides dual-mode OpenXML (.xlsx) generation, parsing, and safe merging:
- Primary mode: native `openpyxl` engine (when available) for 100% Microsoft Excel on Windows / Mac compatibility.
- Fallback mode: zero-dependency pure Python OpenXML builder for minimal / CI environments.

Sheets:
- 00_老板记录: USER-MANAGED protected sheet (never overwritten/cleared).
- 01_项目总览: SYSTEM-MANAGED project-level human-readable dashboard with metadata banner.
- 02_阶段与步骤明细: SYSTEM-MANAGED phase and step breakdown with Done criteria and evidence.
- 03_决策与待办: HYBRID sheet with system decisions and protected user decision/remarks columns.

Follows strict fail-closed preservation: if an existing Excel file cannot be safely parsed,
the original file is never overwritten, returning EXCEL_REFRESH_FAILED_PRESERVED.
"""

from __future__ import annotations

import datetime as dt
import io
import json
import os
import re
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    openpyxl = None
    HAS_OPENPYXL = False


EXCEL_FILE_NAME = "项目进度表_人话版.xlsx"
SHEET_BOSS_LOG = "00_老板记录"
SHEET_PROJECT_OVERVIEW = "01_项目总览"
SHEET_PHASE_STEPS = "02_阶段与步骤明细"
SHEET_DECISIONS = "03_决策与待办"

BOSS_LOG_HEADERS = ["日期", "类型", "我的记录", "以后要做什么", "优先级", "是否已处理"]
PROJECT_OVERVIEW_HEADERS = ["任务ID", "任务名称", "所属阶段", "状态", "完成度%", "负责人/智能体", "一句话人话进度", "当前卡点", "下一步动作", "更新时间"]
PHASE_STEPS_HEADERS = ["任务ID", "阶段序号", "步骤/里程碑名称", "状态", "验收标准 (Done Criteria)", "关联证据文件", "备注"]
DECISION_HEADERS = ["决策项", "背景与影响", "推荐方案", "老板裁决", "老板备注", "裁决时间"]


# --- Fallback Styles XML ---
STYLES_XML = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="7">
    <font><sz val="11"/><name val="Microsoft YaHei"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FFFFFFFF"/><name val="Microsoft YaHei"/><family val="2"/></font>
    <font><i/><sz val="10"/><color rgb="FF555555"/><name val="Microsoft YaHei"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FF155724"/><name val="Microsoft YaHei"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FF004085"/><name val="Microsoft YaHei"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FF856404"/><name val="Microsoft YaHei"/><family val="2"/></font>
    <font><b/><sz val="11"/><color rgb="FF721C24"/><name val="Microsoft YaHei"/><family val="2"/></font>
  </fonts>
  <fills count="8">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF1F4E78"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFD4EDDA"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFCCE5FF"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFFFF3CD"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FFF8D7DA"/><bgColor indexed="64"/></patternFill></fill>
    <fill><patternFill patternType="solid"><fgColor rgb="FF2C3E50"/><bgColor indexed="64"/></patternFill></fill>
  </fills>
  <borders count="2">
    <border><left/><right/><top/><bottom/><diagonal/></border>
    <border>
      <left style="thin"><color rgb="FFD0D7DE"/></left>
      <right style="thin"><color rgb="FFD0D7DE"/></right>
      <top style="thin"><color rgb="FFD0D7DE"/></top>
      <bottom style="thin"><color rgb="FFD0D7DE"/></bottom>
      <diagonal/>
    </border>
  </borders>
  <cellStyleXfs count="1">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>
  </cellStyleXfs>
  <cellXfs count="8">
    <xf numFmtId="0" fontId="0" fillId="0" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="1" fillId="2" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="2" fillId="0" borderId="0" xfId="0" applyFont="1" applyFill="1" applyBorder="0"/>
    <xf numFmtId="0" fontId="3" fillId="3" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="4" fillId="4" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="5" fillId="5" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="6" fillId="6" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
    <xf numFmtId="0" fontId="1" fillId="7" borderId="1" xfId="0" applyFont="1" applyFill="1" applyBorder="1"/>
  </cellXfs>
  <cellStyles count="1">
    <cellStyle name="Normal" xfId="0" builtinId="0"/>
  </cellStyles>
</styleSheet>"""


def get_status_style(status_str: str) -> int:
    """Return matching style index for status strings."""
    s = (status_str or "").strip().upper()
    if any(k in s for k in ("SEALED", "PASS", "DONE", "COMPLETED", "已封板", "已完成", "通过", "全实现")):
        return 3  # Green
    if any(k in s for k in ("DOING", "RUNNING", "IN_PROGRESS", "进行中", "执行中")):
        return 4  # Blue
    if any(k in s for k in ("TODO", "PENDING", "UNSTARTED", "未开始", "待办", "排队")):
        return 5  # Yellow
    if any(k in s for k in ("BLOCKED", "FAIL", "FAILED", "ERROR", "已阻断", "阻断", "失败", "卡点", "异常")):
        return 6  # Red
    return 0


class OpenXMLWorkbookBuilder:
    """Zero-dependency OpenXML (.xlsx) builder with full Microsoft Excel on Windows compatibility."""

    def __init__(self) -> None:
        self.sheets: list[tuple[str, list[list[Any]], list[float]]] = []

    def add_sheet(self, name: str, rows: list[list[Any]], col_widths: list[float] | None = None) -> None:
        self.sheets.append((name, rows, col_widths or []))

    @staticmethod
    def _col_letter(col_idx: int) -> str:
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def build_zip_bytes(self) -> bytes:
        buf = io.BytesIO()
        now_iso = dt.datetime.now(dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
            # 1. [Content_Types].xml
            types_xml = [
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
                '  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
                '  <Default Extension="xml" ContentType="application/xml"/>',
                '  <Override PartName="/docProps/core.xml" ContentType="application/vnd.openxmlformats-package.core-properties+xml"/>',
                '  <Override PartName="/docProps/app.xml" ContentType="application/vnd.openxmlformats-officedocument.extended-properties+xml"/>',
                '  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
                '  <Override PartName="/xl/styles.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml"/>',
            ]
            for idx in range(1, len(self.sheets) + 1):
                types_xml.append(f'  <Override PartName="/xl/worksheets/sheet{idx}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
            types_xml.append('</Types>')
            zf.writestr("[Content_Types].xml", "\n".join(types_xml).encode("utf-8"))

            # 2. _rels/.rels
            rels_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">\n'
                '  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/>\n'
                '  <Relationship Id="rId2" Type="http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties" Target="docProps/core.xml"/>\n'
                '  <Relationship Id="rId3" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/extended-properties" Target="docProps/app.xml"/>\n'
                '</Relationships>'
            )
            zf.writestr("_rels/.rels", rels_xml.encode("utf-8"))

            # 3. docProps/core.xml
            core_xml = (
                f'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                f'<cp:coreProperties xmlns:cp="http://schemas.openxmlformats.org/package/2006/metadata/core-properties" '
                f'xmlns:dc="http://purl.org/dc/elements/1.1/" xmlns:dcterms="http://purl.org/dc/terms/" '
                f'xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">\n'
                f'  <dc:title>项目进度表_人话版</dc:title>\n'
                f'  <dc:creator>Planning-with-Files</dc:creator>\n'
                f'  <cp:lastModifiedBy>Planning-with-Files</cp:lastModifiedBy>\n'
                f'  <dcterms:created xsi:type="dcterms:W3CDTF">{now_iso}</dcterms:created>\n'
                f'  <dcterms:modified xsi:type="dcterms:W3CDTF">{now_iso}</dcterms:modified>\n'
                f'</cp:coreProperties>'
            )
            zf.writestr("docProps/core.xml", core_xml.encode("utf-8"))

            # 4. docProps/app.xml
            sheet_names_xml = "".join(f"<vt:lpstr>{s[0]}</vt:lpstr>" for s in self.sheets)
            app_xml = (
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
                '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument/2006/extended-properties" xmlns:vt="http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes">\n'
                '  <Application>Microsoft Excel</Application>\n'
                '  <DocSecurity>0</DocSecurity>\n'
                '  <ScaleCrop>false</ScaleCrop>\n'
                f'  <HeadingPairs><vt:vector size="2" baseType="variant"><vt:variant><vt:lpstr>Worksheets</vt:lpstr></vt:variant><vt:variant><vt:i4>{len(self.sheets)}</vt:i4></vt:variant></vt:vector></HeadingPairs>\n'
                f'  <TitlesOfParts><vt:vector size="{len(self.sheets)}" baseType="lpstr">{sheet_names_xml}</vt:vector></TitlesOfParts>\n'
                '  <Company></Company>\n'
                '  <LinksUpToDate>false</LinksUpToDate>\n'
                '  <SharedDoc>false</SharedDoc>\n'
                '  <HyperlinksChanged>false</HyperlinksChanged>\n'
                '  <AppVersion>16.0300</AppVersion>\n'
                '</Properties>'
            )
            zf.writestr("docProps/app.xml", app_xml.encode("utf-8"))

            # 5. xl/_rels/workbook.xml.rels
            wb_rels = [
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">',
            ]
            for idx in range(1, len(self.sheets) + 1):
                wb_rels.append(f'  <Relationship Id="rId{idx}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet{idx}.xml"/>')
            styles_rid = len(self.sheets) + 1
            wb_rels.append(f'  <Relationship Id="rId{styles_rid}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/styles" Target="styles.xml"/>')
            wb_rels.append('</Relationships>')
            zf.writestr("xl/_rels/workbook.xml.rels", "\n".join(wb_rels).encode("utf-8"))

            # 6. xl/workbook.xml
            wb_xml = [
                '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
                '  <workbookPr/>',
                '  <sheets>',
            ]
            for idx, (name, _, _) in enumerate(self.sheets, start=1):
                clean_name = name.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace('"', "&quot;")
                wb_xml.append(f'    <sheet name="{clean_name}" sheetId="{idx}" r:id="rId{idx}"/>')
            wb_xml.append('  </sheets>')
            wb_xml.append('  <calcPr calcId="124519" fullCalcOnLoad="1"/>')
            wb_xml.append('</workbook>')
            zf.writestr("xl/workbook.xml", "\n".join(wb_xml).encode("utf-8"))

            # 7. Worksheets
            for sheet_idx, (name, rows, col_widths) in enumerate(self.sheets, start=1):
                max_row = max(1, len(rows))
                max_col = max(1, max((len(r) for r in rows), default=1))
                dim_ref = f"A1:{self._col_letter(max_col)}{max_row}"

                ws_xml = [
                    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                    '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
                    f'  <dimension ref="{dim_ref}"/>',
                    '  <sheetViews><sheetView workbookViewId="0"><selection activeCell="A1" sqref="A1"/></sheetView></sheetViews>',
                    '  <sheetFormatPr defaultRowHeight="15"/>',
                ]
                if col_widths:
                    ws_xml.append('  <cols>')
                    for col_i, width in enumerate(col_widths, start=1):
                        ws_xml.append(f'    <col min="{col_i}" max="{col_i}" width="{width}" customWidth="1"/>')
                    ws_xml.append('  </cols>')
                ws_xml.append('  <sheetData>')
                for row_idx, row_cells in enumerate(rows, start=1):
                    if not row_cells:
                        ws_xml.append(f'    <row r="{row_idx}"/>')
                        continue
                    ws_xml.append(f'    <row r="{row_idx}">')
                    for col_idx, cell_data in enumerate(row_cells, start=1):
                        c_ref = f"{self._col_letter(col_idx)}{row_idx}"
                        cell_val = cell_data
                        style_idx = 0
                        if isinstance(cell_data, dict):
                            cell_val = cell_data.get("val", "")
                            style_idx = cell_data.get("style", 0)

                        str_val = str(cell_val) if cell_val is not None else ""
                        escaped_val = (
                            str_val.replace("&", "&amp;")
                            .replace("<", "&lt;")
                            .replace(">", "&gt;")
                        )
                        escaped_val = "".join(ch for ch in escaped_val if ord(ch) >= 32 or ch in "\t\n\r")

                        if style_idx > 0:
                            ws_xml.append(f'      <c r="{c_ref}" t="inlineStr" s="{style_idx}"><is><t>{escaped_val}</t></is></c>')
                        else:
                            ws_xml.append(f'      <c r="{c_ref}" t="inlineStr"><is><t>{escaped_val}</t></is></c>')
                    ws_xml.append('    </row>')
                ws_xml.append('  </sheetData>')
                ws_xml.append('  <pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>')
                ws_xml.append('</worksheet>')
                zf.writestr(f"xl/worksheets/sheet{sheet_idx}.xml", "\n".join(ws_xml).encode("utf-8"))

            # 8. xl/styles.xml
            zf.writestr("xl/styles.xml", STYLES_XML.encode("utf-8"))

        return buf.getvalue()


def build_workbook_openpyxl(sheets_data: list[tuple[str, list[list[Any]], list[float]]]) -> bytes:
    """Build spreadsheet using openpyxl for 100% native Microsoft Excel on Windows fidelity."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    font_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    font_boss_header = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Microsoft YaHei", size=10)
    font_meta = Font(name="Microsoft YaHei", size=10, italic=True, color="555555")

    fill_header = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    fill_boss = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")

    status_styles = {
        "SEALED": (Font(name="Microsoft YaHei", size=10, bold=True, color="155724"), PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")),
        "PASS": (Font(name="Microsoft YaHei", size=10, bold=True, color="155724"), PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")),
        "DOING": (Font(name="Microsoft YaHei", size=10, bold=True, color="004085"), PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")),
        "TODO": (Font(name="Microsoft YaHei", size=10, bold=True, color="856404"), PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")),
        "BLOCKED": (Font(name="Microsoft YaHei", size=10, bold=True, color="721C24"), PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")),
    }

    thin_border = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    for title, rows, col_widths in sheets_data:
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        is_boss_sheet = (title == SHEET_BOSS_LOG)
        is_overview_sheet = (title == SHEET_PROJECT_OVERVIEW)

        for r_idx, row in enumerate(rows, start=1):
            if not row:
                continue
            is_header_row = (is_boss_sheet and r_idx == 1) or (is_overview_sheet and r_idx == 6) or (not is_boss_sheet and not is_overview_sheet and r_idx == 1)
            is_meta_row = is_overview_sheet and (1 <= r_idx <= 4)

            for c_idx, cell_data in enumerate(row, start=1):
                cell_val = cell_data
                style_override = None
                if isinstance(cell_data, dict):
                    cell_val = cell_data.get("val", "")
                    style_override = cell_data.get("style", None)

                c = ws.cell(row=r_idx, column=c_idx, value=cell_val)
                c.font = font_body
                c.border = thin_border
                c.alignment = Alignment(vertical="center")

                if is_meta_row:
                    c.font = font_meta
                    c.border = Border()
                elif is_header_row:
                    c.font = font_boss_header if is_boss_sheet else font_header
                    c.fill = fill_boss if is_boss_sheet else fill_header
                    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                elif style_override in (3, 4, 5, 6):
                    key = {3: "PASS", 4: "DOING", 5: "TODO", 6: "BLOCKED"}.get(style_override)
                    if key and key in status_styles:
                        c.font, c.fill = status_styles[key]
                        c.alignment = Alignment(horizontal="center", vertical="center")

        for c_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(c_idx)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_existing_xlsx(xlsx_path: Path | str) -> dict[str, list[list[str]]]:
    """Parse existing .xlsx file using openpyxl (if available) or standard library."""
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        return {}

    if HAS_OPENPYXL:
        try:
            wb = openpyxl.load_workbook(xlsx_path, data_only=True)
            sheets_data: dict[str, list[list[str]]] = {}
            for name in wb.sheetnames:
                ws = wb[name]
                rows: list[list[str]] = []
                for r in ws.iter_rows(values_only=True):
                    row_vals = [str(v) if v is not None else "" for v in r]
                    if name == SHEET_DECISIONS and len(row_vals) < len(DECISION_HEADERS):
                        row_vals.extend([""] * (len(DECISION_HEADERS) - len(row_vals)))
                    elif name == SHEET_BOSS_LOG and len(row_vals) < len(BOSS_LOG_HEADERS):
                        row_vals.extend([""] * (len(BOSS_LOG_HEADERS) - len(row_vals)))
                    rows.append(row_vals)
                sheets_data[name] = rows
            return sheets_data
        except Exception:
            # Fallback to standard library parser if openpyxl fails on legacy/malformed files
            pass

    with zipfile.ZipFile(xlsx_path, "r") as zf:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            tree = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            ns = {"ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
            for si in tree.findall(".//ns:si", ns):
                text_parts = [t.text or "" for t in si.findall(".//ns:t", ns)]
                shared_strings.append("".join(text_parts))

        wb_tree = ET.fromstring(zf.read("xl/workbook.xml"))
        ns = {
            "ns": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
            "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        }

        wb_rels_tree = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
        rel_ns = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
        rel_map = {}
        for rel in wb_rels_tree.findall("rel:Relationship", rel_ns):
            rel_map[rel.attrib["Id"]] = rel.attrib["Target"]

        sheets_data = {}
        for sheet in wb_tree.findall(".//ns:sheet", ns):
            name = sheet.attrib["name"]
            rid = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
            if not rid and "r:id" in sheet.attrib:
                rid = sheet.attrib["r:id"]
            target = rel_map.get(rid, "")
            if not target:
                continue
            if not target.startswith("xl/"):
                target = f"xl/{target.lstrip('/')}"
            if target not in zf.namelist():
                continue

            ws_tree = ET.fromstring(zf.read(target))
            rows = []
            for row_el in ws_tree.findall(".//ns:row", ns):
                row_dict: dict[int, str] = {}
                for c_el in row_el.findall("ns:c", ns):
                    r_attr = c_el.attrib.get("r", "")
                    col_letters = "".join(filter(str.isalpha, r_attr))
                    if not col_letters:
                        continue
                    col_idx = 0
                    for char in col_letters:
                        col_idx = col_idx * 26 + (ord(char.upper()) - 64)

                    t_attr = c_el.attrib.get("t", "")
                    val = ""
                    if t_attr == "s":
                        v_el = c_el.find("ns:v", ns)
                        if v_el is not None and v_el.text:
                            try:
                                s_idx = int(v_el.text)
                                if 0 <= s_idx < len(shared_strings):
                                    val = shared_strings[s_idx]
                            except ValueError:
                                pass
                    elif t_attr == "inlineStr":
                        t_el = c_el.find(".//ns:t", ns)
                        if t_el is not None and t_el.text:
                            val = t_el.text
                    else:
                        v_el = c_el.find("ns:v", ns)
                        if v_el is not None and v_el.text:
                            val = v_el.text

                    row_dict[col_idx] = val

                if row_dict:
                    max_c = max(row_dict.keys())
                    row_list = [row_dict.get(c_i, "") for c_i in range(1, max_c + 1)]
                    if name == SHEET_DECISIONS and len(row_list) < len(DECISION_HEADERS):
                        row_list.extend([""] * (len(DECISION_HEADERS) - len(row_list)))
                    elif name == SHEET_BOSS_LOG and len(row_list) < len(BOSS_LOG_HEADERS):
                        row_list.extend([""] * (len(BOSS_LOG_HEADERS) - len(row_list)))
                    rows.append(row_list)
                else:
                    rows.append([])
            sheets_data[name] = rows

        return sheets_data


def _parse_frontmatter(content: str) -> dict[str, str]:
    """Parse YAML frontmatter from markdown file."""
    if not content.startswith("---"):
        return {}
    parts = content.split("---", 2)
    if len(parts) < 3:
        return {}
    fm_lines = parts[1].strip().split("\n")
    data: dict[str, str] = {}
    for line in fm_lines:
        line = line.strip()
        if ":" in line and not line.startswith("#"):
            k, v = line.split(":", 1)
            data[k.strip()] = v.strip().strip('"').strip("'")
    return data


def extract_project_data(planning_root: Path) -> dict[str, Any]:
    """Extract structured data for sheets 01, 02, and 03 from planning documents."""
    tasks: list[dict[str, Any]] = []
    phase_steps: list[dict[str, Any]] = []
    decisions: list[dict[str, Any]] = []

    # Find task directories
    task_dirs: list[Path] = []
    if planning_root.is_dir():
        for item in sorted(planning_root.iterdir()):
            if item.is_dir() and not item.name.startswith((".", "_")) and item.name not in ("ADR", "evidence", "reports"):
                if (item / "1_master_plan.md").exists() or (item / "task_plan.md").exists() or (item / "3_status_update.md").exists():
                    task_dirs.append(item)

    # Fallback: if planning_root itself contains 1_master_plan.md
    if not task_dirs and (planning_root / "1_master_plan.md").exists():
        task_dirs.append(planning_root)

    # Check project-level index for global status
    global_status = ""
    index_file = planning_root / "00_PROJECT_INDEX.md"
    if not index_file.exists() and planning_root.parent:
        index_file = planning_root.parent / "00_PROJECT_INDEX.md"
    if index_file.exists():
        idx_text = index_file.read_text(encoding="utf-8", errors="ignore")
        if "TASK_STATUS=COMPLETE" in idx_text or "SYSTEM_STATUS=EFFECTIVE" in idx_text:
            global_status = "SEALED"

    for t_dir in task_dirs:
        task_id = t_dir.name
        master_plan_file = t_dir / "1_master_plan.md"
        status_file = t_dir / "3_status_update.md"
        task_plan_file = t_dir / "task_plan.md"

        fm: dict[str, str] = {}
        title = task_id
        summary = ""
        owner = "mac-codex"
        status = global_status or "TODO"
        phase = "V1"
        updated = dt.date.today().strftime("%Y-%m-%d")
        total_cbs = 0
        done_cbs = 0

        if master_plan_file.exists():
            text = master_plan_file.read_text(encoding="utf-8", errors="ignore")
            fm = _parse_frontmatter(text)
            title = fm.get("title", title)
            summary = fm.get("summary", "")
            owner = fm.get("owner", owner)
            if "status" in fm:
                status = fm["status"]
            phase = fm.get("phase", phase)
            updated = fm.get("updated", updated)

            # Extract title from markdown H1 if not set in frontmatter
            if title == task_id:
                t_match = re.search(r"^#\s*([^#\n]+?)(?:\s*项目主计划|\s*主计划|\s*Master Plan)?\s*$", text, re.MULTILINE)
                if t_match:
                    title = t_match.group(1).strip()

            # Extract summary from first section heading or blockquote
            if not summary:
                first_sec = re.search(r"##\s*[^#\n]+\n+([^#\n>][^\n]+)", text)
                if first_sec:
                    summary = first_sec.group(1).strip()
                else:
                    quote_match = re.search(r">\s*(?:实施结果|当前归档状态|当前状态)[：:]\s*`?([^`\n]+)`?", text)
                    if quote_match:
                        summary = quote_match.group(1).strip()

            if "TASK_STATUS=COMPLETE" in text:
                status = "SEALED"

            # 1. Extract checkbox steps from Done Criteria
            cb_matches = re.findall(r"-\s*\[([ xX])\]\s*(.+)", text)
            if cb_matches:
                total_cbs = len(cb_matches)
                done_cbs = sum(1 for m, _ in cb_matches if m.lower() == "x")
                for idx, (mark, step_text) in enumerate(cb_matches, start=1):
                    is_done = mark.lower() == "x"
                    step_clean = re.sub(r"[*_`]", "", step_text).strip()
                    phase_steps.append({
                        "task_id": task_id,
                        "phase_no": f"P{idx}",
                        "step_name": step_clean,
                        "status": "PASS" if is_done else "TODO",
                        "done_criteria": "Done Criteria 验收",
                        "evidence_ref": f"{task_id}/evidence",
                        "notes": "",
                    })

            # 2. Extract numbered phase steps from "## 阶段与 Done Criteria"
            if not cb_matches:
                phase_section = re.search(r"##\s*阶段与\s*Done Criteria\s*\n(.*?)(?=\n##|\Z)", text, re.DOTALL)
                if phase_section:
                    lines = phase_section.group(1).strip().split("\n")
                    for line in lines:
                        line = line.strip()
                        m = re.match(r"^(\d+)\.\s*(.+)", line)
                        if m:
                            step_text = m.group(2).strip()
                            phase_steps.append({
                                "task_id": task_id,
                                "phase_no": f"P{m.group(1)}",
                                "step_name": step_text,
                                "status": "PASS" if status.upper() in ("SEALED", "PASS") else "TODO",
                                "done_criteria": "完成标准校验",
                                "evidence_ref": f"{task_id}/evidence",
                                "notes": "",
                            })

        if status_file.exists():
            st_text = status_file.read_text(encoding="utf-8", errors="ignore")
            st_fm = _parse_frontmatter(st_text)
            if "status" in st_fm:
                status = st_fm["status"]
            if "updated" in st_fm:
                updated = st_fm["updated"]

            # Parse markdown gates table | Gate | 状态 | 证据 |
            table_matches = re.findall(r"\|\s*([^|\n]+)\s*\|\s*([^|\n]+)\s*\|\s*([^|\n]+)\s*\|", st_text)
            if len(table_matches) > 1:
                for row in table_matches[1:]:
                    gate_name, g_status, g_evidence = [col.strip() for col in row]
                    if gate_name.startswith("-") or gate_name in ("Gate", "检查项", "门禁"):
                        continue
                    phase_steps.append({
                        "task_id": task_id,
                        "phase_no": "Gate",
                        "step_name": gate_name,
                        "status": g_status,
                        "done_criteria": "Gate Validation",
                        "evidence_ref": g_evidence,
                        "notes": "",
                    })

            # Extract decisions from ## 【老板待裁决区】
            boss_dec_match = re.search(r"##\s*【?老板待裁决(?:区)?】?\s*\n(.*?)(?=\n##|\Z)", st_text, re.DOTALL)
            if boss_dec_match:
                lines = [l.strip() for l in boss_dec_match.group(1).split("\n") if l.strip().startswith("-")]
                for line in lines[:5]:
                    content = re.sub(r"^-\s*\[.*?\]\s*", "", line).strip()
                    if content:
                        d_item = re.split(r"[：:，,。]", content)[0].strip()
                        decisions.append({
                            "decision_item": d_item[:40],
                            "context": content,
                            "recommendation": "待老板裁决",
                            "boss_decision": "",
                            "boss_notes": "",
                            "decision_time": "",
                        })

            # Extract decisions from ## D｜当前有效决策
            valid_dec_match = re.search(r"##\s*(?:D[｜|])?当前有效决策\s*\n(.*?)(?=\n##|\Z)", st_text, re.DOTALL)
            if valid_dec_match:
                lines = [l.strip() for l in valid_dec_match.group(1).split("\n") if l.strip().startswith("-")]
                for line in lines[:5]:
                    content = re.sub(r"^-\s*\[.*?\]\s*", "", line).strip()
                    if content:
                        d_item = re.split(r"[：:，,。]", content)[0].strip()
                        decisions.append({
                            "decision_item": d_item[:40],
                            "context": content,
                            "recommendation": "已生效执行",
                            "boss_decision": "已执行",
                            "boss_notes": "系统已落实",
                            "decision_time": "",
                        })

        # Progress percentage & human summary details
        if status.upper() in ("SEALED", "PASS", "COMPLETED"):
            progress_pct = "100%"
            blockers = "无"
            next_step = "已封板归档 / 日常维护与按需修复"
        elif status.upper() in ("DOING", "RUNNING"):
            if total_cbs > 0:
                progress_pct = f"{int((done_cbs / total_cbs) * 100)}%"
            else:
                progress_pct = "50%"
            blockers = "无"
            next_step = "继续推进下一阶段"
        elif status.upper() in ("BLOCKED", "FAIL", "FAILED"):
            progress_pct = "0%"
            blockers = "存在卡点阻断"
            next_step = "定位并消除阻断项"
        else:
            progress_pct = "0%"
            blockers = "无"
            next_step = "开始执行"

        tasks.append({
            "task_id": task_id,
            "task_name": title,
            "phase": phase,
            "status": status.upper(),
            "progress_pct": progress_pct,
            "owner": owner,
            "summary": summary or f"任务 {task_id} 正在推进",
            "blockers": blockers,
            "next_step": next_step,
            "updated": updated,
        })

    # Default fallback task if no tasks found
    if not tasks:
        tasks.append({
            "task_id": "PROJECT-ROOT",
            "task_name": "项目初始化",
            "phase": "INIT",
            "status": "DOING",
            "progress_pct": "30%",
            "owner": "mac-codex",
            "summary": "项目规划与治理地基已建立",
            "blockers": "无",
            "next_step": "制定任务分期与验收标准",
            "updated": dt.date.today().strftime("%Y-%m-%d"),
        })

    # Default decision items if none extracted
    if not decisions:
        decisions.append({
            "decision_item": "DEC-001: 治理档位选择 (L0-L3)",
            "context": "根据任务复杂度确定规划深度与门禁级别",
            "recommendation": "常规项目推荐 L2 (STANDARD)，高安全要求推荐 L3 (STRICT)",
            "boss_decision": "",
            "boss_notes": "",
            "decision_time": "",
        })

    return {
        "tasks": tasks,
        "phase_steps": phase_steps,
        "decisions": decisions,
    }


def generate_progress_excel(
    project_root: Path | str,
    output_path: Path | str | None = None,
) -> tuple[bool, str]:
    """Generate or update the human-readable project progress Excel sheet.

    Safely preserves USER-MANAGED sheet 00_老板记录 and user columns in 03_决策与待办.
    Follows fail-closed: never corrupts or overwrites original file on parsing failure.

    Returns:
        (True, "EXCEL_REFRESH_SUCCESS: <path>") or (False, "EXCEL_REFRESH_FAILED_PRESERVED: <reason>")
    """
    project_root = Path(project_root).resolve()
    planning_root = project_root / "00.项目规划与治理"
    if not planning_root.exists():
        planning_root = project_root

    if output_path is None:
        target_file = planning_root / EXCEL_FILE_NAME
    else:
        target_file = Path(output_path).resolve()

    # 1. Check if target file exists and extract user-managed data
    preserved_boss_rows: list[list[str]] = []
    preserved_decision_map: dict[str, dict[str, str]] = {}
    is_existing = target_file.exists()

    if is_existing:
        try:
            old_sheets = parse_existing_xlsx(target_file)
            # Extract 00_老板记录 (all rows)
            if SHEET_BOSS_LOG in old_sheets:
                preserved_boss_rows = old_sheets[SHEET_BOSS_LOG]

            # Extract 03_决策与待办 user columns
            if SHEET_DECISIONS in old_sheets:
                d_rows = old_sheets[SHEET_DECISIONS]
                # Header is row 0: 决策项 | 背景与影响 | 推荐方案 | 老板裁决 | 老板备注 | 裁决时间
                for row in d_rows[1:]:
                    if len(row) > 0 and row[0].strip():
                        d_item = row[0].strip()
                        boss_decision = row[3].strip() if len(row) > 3 else ""
                        boss_notes = row[4].strip() if len(row) > 4 else ""
                        decision_time = row[5].strip() if len(row) > 5 else ""
                        if boss_decision or boss_notes or decision_time:
                            preserved_decision_map[d_item] = {
                                "boss_decision": boss_decision,
                                "boss_notes": boss_notes,
                                "decision_time": decision_time,
                            }
        except Exception as exc:
            # Fail-closed: do not overwrite damaged or unparseable existing file
            return (False, f"EXCEL_REFRESH_FAILED_PRESERVED: failed to parse existing workbook: {exc}")

    # 2. Extract current project data from Markdown sources
    try:
        data = extract_project_data(planning_root)
    except Exception as exc:
        return (False, f"EXCEL_REFRESH_FAILED_PRESERVED: failed to extract project markdown data: {exc}")

    # 3. Build sheet rows
    # --- Sheet 1: 00_老板记录 ---
    boss_sheet_rows: list[list[Any]] = []
    if preserved_boss_rows:
        # Preserve user rows exactly
        for r_idx, r in enumerate(preserved_boss_rows):
            if r_idx == 0:
                boss_sheet_rows.append([{"val": c, "style": 7} for c in r])
            else:
                boss_sheet_rows.append(r)
    else:
        # Create empty template
        boss_sheet_rows.append([{"val": h, "style": 7} for h in BOSS_LOG_HEADERS])

    # --- Sheet 2: 01_项目总览 ---
    now_str = dt.datetime.now().strftime("%Y-%m-%d %H:%M")
    overview_rows: list[list[Any]] = [
        [{"val": "权威来源：Planning Files (Markdown / Checkpoints)", "style": 2}],
        [{"val": f"最近同步：{now_str}", "style": 2}],
        [{"val": "Excel状态：已同步", "style": 2}],
        [{"val": "说明：本表为面向老板的可视化进度视图，不作为正式状态源。", "style": 2}],
        [],  # Blank separator row
        [{"val": h, "style": 1} for h in PROJECT_OVERVIEW_HEADERS],
    ]
    for t in data["tasks"]:
        st_style = get_status_style(t["status"])
        overview_rows.append([
            t["task_id"],
            t["task_name"],
            t["phase"],
            {"val": t["status"], "style": st_style},
            t["progress_pct"],
            t["owner"],
            t["summary"],
            t["blockers"],
            t["next_step"],
            t["updated"],
        ])

    # --- Sheet 3: 02_阶段与步骤明细 ---
    step_rows: list[list[Any]] = [
        [{"val": h, "style": 1} for h in PHASE_STEPS_HEADERS]
    ]
    for s in data["phase_steps"]:
        st_style = get_status_style(s["status"])
        step_rows.append([
            s["task_id"],
            s["phase_no"],
            s["step_name"],
            {"val": s["status"], "style": st_style},
            s["done_criteria"],
            s["evidence_ref"],
            s["notes"],
        ])

    # --- Sheet 4: 03_决策与待办 ---
    dec_rows: list[list[Any]] = [
        [{"val": h, "style": 1} for h in DECISION_HEADERS]
    ]
    for d in data["decisions"]:
        d_item = d["decision_item"]
        # Merge preserved user columns if available
        user_vals = preserved_decision_map.get(d_item, {})
        boss_decision = user_vals.get("boss_decision", d["boss_decision"])
        boss_notes = user_vals.get("boss_notes", d["boss_notes"])
        dec_time = user_vals.get("decision_time", d["decision_time"])

        dec_rows.append([
            d_item,
            d["context"],
            d["recommendation"],
            boss_decision,
            boss_notes,
            dec_time,
        ])

    sheets_to_build = [
        (SHEET_BOSS_LOG, boss_sheet_rows, [15, 12, 40, 35, 10, 12]),
        (SHEET_PROJECT_OVERVIEW, overview_rows, [22, 28, 14, 12, 10, 15, 38, 18, 22, 14]),
        (SHEET_PHASE_STEPS, step_rows, [20, 12, 35, 12, 30, 25, 20]),
        (SHEET_DECISIONS, dec_rows, [30, 35, 35, 18, 25, 15]),
    ]

    # 4. Generate byte stream and write atomically via temp file
    try:
        if HAS_OPENPYXL:
            xlsx_bytes = build_workbook_openpyxl(sheets_to_build)
        else:
            builder = OpenXMLWorkbookBuilder()
            for s_name, s_rows, s_widths in sheets_to_build:
                builder.add_sheet(s_name, s_rows, s_widths)
            xlsx_bytes = builder.build_zip_bytes()

        target_file.parent.mkdir(parents=True, exist_ok=True)

        with tempfile.NamedTemporaryFile(dir=target_file.parent, delete=False, suffix=".tmp") as tf:
            tf.write(xlsx_bytes)
            temp_path = Path(tf.name)

        # Atomic replace
        os.replace(temp_path, target_file)
        return (True, f"EXCEL_REFRESH_SUCCESS: {target_file}")
    except Exception as exc:
        if "temp_path" in locals() and temp_path.exists():
            temp_path.unlink(missing_ok=True)
        return (False, f"EXCEL_REFRESH_FAILED_PRESERVED: failed to generate or write excel: {exc}")
